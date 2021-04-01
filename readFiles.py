"""To read xls and xlsx files for name/date to organize files"""


from openpyxl import load_workbook


def read_expense(expense_sheet):
    """Reads expense file, returns the name and date on report"""

    wb = load_workbook(expense_sheet, data_only=True)
    sheet = wb['Expenses']

    tech_name = sheet['B5'].value
    expense_date = str(sheet['B3'].value).split(' ')[0]

    return tech_name, expense_date


def read_bonus(bonus_sheet):
    """Reads bonus file, returns the name and date on report"""

    wb = load_workbook(bonus_sheet, data_only=True)
    sheet = wb['Bonuses']

    tech_name = sheet['C3'].value
    bonus_date = str(sheet['C4'].value).split(' ')[0]

    return tech_name, bonus_date


def expense_variables(e_sheet):
    """Reads expense fine and returns the date and expense total"""

    wb = load_workbook(e_sheet, data_only=True)
    sheet = wb['Expenses']

    expense_date = str(sheet['B3'].value).split(' ')[0]
    expense_total = sheet['J30'].value

    return expense_date, expense_total


def bonus_variables(b_sheet):
    """Reads bonus sheet and returns date, travel, oncall, and call case sums"""

    wb = load_workbook(b_sheet, data_only=True)
    sheet = wb['Bonuses']

    bonus_date = str(sheet['C4'].value).split(' ')[0]
    travel = sheet['B19'].value
    on_call = sheet['D19'].value
    call_cases = sheet['F19'].value

    return bonus_date, travel, on_call, call_cases
